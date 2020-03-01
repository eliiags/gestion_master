#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
@author: Elianni Aguero, Angelica Guerrero, Cynthia Quintana
"""


# Comprobar si el fichero existe
import os
# Para guardar los datos en EXCEL 
from openpyxl import Workbook
import openpyxl
# Graficos
#from IPython import get_ipython
#get_ipython().run_line_magic('matplotlib', 'inline')
import matplotlib.pyplot as grafico
import operator

import Utiles
import ModoJuego


# Variables globales
# Ruta y nombre del fichero
FICHERO_PARTIDAS = '/home/eli/MASTER/DATOS/partidas.xlsx'
###   Diccionario con informacion del juego
DatosJuego = {'N1':20, 
              'N2':12,
              'N3':5,
              'Int1':[1, 1000],
              'Int2':[1, 10000], 
              'Int3':[-100, 100]}



class Menu():
    def __init__(self):
        # Se guardarán los datos del jugador para actualizarlos
        self._datosJugadores = ""

    # Mostramos el menu del modo del juego
    def muestraMenuModoJuego(self):
        print("\n¿Cómo quieres jugar?") 
        print(" 1. Partida modo solitario")
        print(" 2. Partida 2 Jugadores") 
        print(" 3. Estadística")
        print(" 4. Salir \n")

    # Mostramos el menu del nivel del juego (F, M, D)
    def muestraMenuNivelJuego(self):
        print("selecciona el nivel del juego")
        print(" 1. Fácil ("   + str(DatosJuego['N1']) + " intentos)")
        print(" 2. Medio ("   + str(DatosJuego['N2']) + " intentos)")
        print(" 3. Difícil (" + str(DatosJuego['N3']) + " intentos) \n")
        
    # Intervalo en el que quiere adivinar el numero
    def muestraIntervalos(self):
        print("Ahora selecciona el intervalo")
        print(" 1. Adivinar un número entre " + str(DatosJuego['Int1'][0]) + " y " 
                                             + str(DatosJuego['Int1'][1]))
        print(" 2. Adivinar un número entre " + str(DatosJuego['Int2'][0]) + " y " 
                                             + str(DatosJuego['Int2'][1]))
        print(" 3. Adivinar un número entre " + str(DatosJuego['Int3'][0]) + " y " 
                                             + str(DatosJuego['Int3'][1]) + "\n")
    
    def muestraMenuEstadisticas(self):
        print("\nSelecciona qué estadísticas quieres ver")
        print(" 1. Partidas Ganadas / Perdidas")
        print(" 2. Partidas ganadas por niveles")
        print(" 3. Ranking de jugadores")
        print(" 4. Volver al menu principal\n")
    
    
    # Comprobamos que la opcion seleccionada es valida
    # La variable "menu" contiene el indicador de si es menu de 
    # - 1 modo o 
    # - 2 nivel
    def opcionSeleccionada(self, n1, n2):
        opcion = input("Selecciona una de las opciones anteriores: ")
        # Se comprueba que el numero introducido es correcto. Es decir, si el valor 
        # introducido es un numero y si, ademas, esta en el intervalo correcto.
        # Además se le añade el parámetro False o True para saber si es un número 
        # oculto o no
        # Esta funcion compruebaNumero devuelve el numero correcto
        return Utiles.compruebaNumero(n1, n2, opcion, False)


    def menu(self):
        # Saludo de bienvenida
        print(Utiles.getColor('V') + "Bienvenido a 'ADIVINA EL NÚMERO'" + Utiles.getColor('B'))
        
        self._datosJugadores = self.extraerBBDD()
        
        while(True):
            ### Muestra el menu del Modo de Juego ###
            self.muestraMenuModoJuego()
            opcionModo = self.opcionSeleccionada(1, 4)
            
            # Opcion Estadistica
            if (opcionModo == 3):
                self.mostrarEstadisticas()
            # Salir del juego
            elif (opcionModo == 4):     
                ## Se actualiza la BBDD
                self.actualizaBBDD()
                print("Hasta la próxima! :)")
                break
            # Modo jugar
            else:
                print("Muy bien! Has seleccionado la opción ", end="")
                if (opcionModo == 1):
                    print("Modo Solitario. Ahora ", end="")
                else:
                    print("Dos Jugadores. Ahora ", end="")
            
            
                # Muestra el menu del Nivel de Juego
                self.muestraMenuNivelJuego()
                opcionNivel = self.opcionSeleccionada(1, 3)
            
                # Muestra el menu de Intervalos
                self.muestraIntervalos()
                opcionIntervalo = self.opcionSeleccionada(1, 3)
            
                # The game is on :)
                self.empiezaElJuego(opcionModo, opcionNivel, opcionIntervalo)
            
            
            # Volvemos a establecer los valores a predeterminados
            self._existe = False
            self._pos    = 0
            self._listaPuntos = ""



    # Desarrollo del juego
    def empiezaElJuego(self, opcionModo, opcionNivel, opcionIntervalo):
        
        juego = ""
        # Segun la opcion crea el tipo de juego
        if (opcionModo == 1):
            # MODO SOLITARIO 
            # Se le pasa el nivel seleccionado por el usuario
            juego = ModoJuego.Solitario(int(DatosJuego["Int" + str(opcionIntervalo)][0]),
                                        int(DatosJuego["Int" + str(opcionIntervalo)][1]),
                                        int(DatosJuego["N" + str(opcionNivel)]))
        else:
            # MODO MULTIJUGADOR
            # Se le pasa el nivel seleccionado por el usuario
            juego = ModoJuego.MultiJugador(int(DatosJuego["Int" + str(opcionIntervalo)][0]),
                                           int(DatosJuego["Int" + str(opcionIntervalo)][1]),
                                           int(DatosJuego["N" + str(opcionNivel)]))
        
        
        # the game is on
        juego.jugar()
        
        nombres = juego.getInfoJugadores()
        
        # Comprobamos si el nombre del jugador esta en la BBDD
        # Si no creamos una nueva observacion para el mismo 
        for n in nombres:
            if (n not in self._datosJugadores):
                self._datosJugadores[n] = [0, 0, 0, 0, 0]
        
        
        # Se actualiza la informacion del jugador sea nuevo o no
        # SI HA GANADO
        if (juego.haGanado()):
            # Actualizamos la partida ganada
            self._datosJugadores[nombres[0]][0] += 1
            if (opcionModo == 2):
                # Se actualiza al JUGADOR Nº1 como partida perdida
                self._datosJugadores[nombres[1]][1] += 1
            
            self._datosJugadores[nombres[0]][opcionNivel+1] +=1
          
        # SI HA PERDIDO
        else:
            # Actualizamos la partida perdida
            self._datosJugadores[nombres[0]][1] += 1
            if (opcionModo == 2):
                # Se actualiza al JUGADOR Nº1 como partida ganada
                self._datosJugadores[nombres[1]][1] += 1
                self._datosJugadores[nombres[1]][opcionNivel+1] +=1
            

    # Accede al fichero excel y comprueba el nombre del jugador
    # Si el jugador existe recoge sus datos
    # Si no existe crea uno nuevo y pone toda la info a 0 
    def extraerBBDD(self):
      
        fichero = ""
        datosJugadores = {}
        nombre  = ""
        # Comprueba si el fichero ya existe
        if os.path.exists(FICHERO_PARTIDAS):
            # Abrimos el fichero para comprobar si el jugador ya existe
            fichero = openpyxl.load_workbook(FICHERO_PARTIDAS)
            hoja    = fichero.active
            if (hoja.cell(1, 1).value is None):
                return datosJugadores
            # Guardamos toda la info del excel en el atributo datosJugadores
            # Se guardarán en un diccionario cuya clave sera el nombre del jugador
            for i in range(1, hoja.max_row+1):
                listaInfo = []
                nombre = hoja.cell(i, 1).value
                for j in range(2, hoja.max_column+1):
                    listaInfo.append(hoja.cell(i,j).value)
                datosJugadores[nombre] = listaInfo
            
        # Si el fichero no existe 
        else:
            # lo creamos
            wb = Workbook()
            # Lo guardamos
            wb.save(FICHERO_PARTIDAS)
        
        return datosJugadores
        

    # Se actualiza la info de la partida        
    def actualizaBBDD(self):
        
        # Abrimos el fichero 
        fichero = openpyxl.load_workbook(FICHERO_PARTIDAS)
        hoja    = fichero.active
        
        nombres = list(self._datosJugadores.keys())
        
        # Guardamos toda la info del excel en el atributo datosJugadores
        for i in range(1, len(self._datosJugadores)+1):
            #print("Ite i: " + str(i))
            for j in range(1, len(self._datosJugadores[nombres[i-1]])+2):
                #print("Ite j: " + str(j))
                if (j == 1):
                    hoja.cell(i, j).value = nombres[i-1]
                else: 
                    hoja.cell(i, j).value = self._datosJugadores[nombres[i-1]][j-2]
        
        fichero.save(FICHERO_PARTIDAS)

    
    # Muestra un gráfico con las estadísticas del usuario
    def mostrarEstadisticas(self):
        
        nombre = input("Escribe tu nombre: ").upper()
        
        if (nombre not in self._datosJugadores):
            print("\nNunca has jugado a ADIVINA EL NUMERO, " + 
              "por lo que no tenemos estadísticas para mostrarte. " +
              "¿Te animas a jugar una partida?\n ")
            return ""
        
        print(Utiles.getColor('A') + "\nEstadísticas de: " + nombre + Utiles.getColor('B'))

        while(True):
            self.muestraMenuEstadisticas()
            opcionEstadistica = self.opcionSeleccionada(1, 4)

            if (opcionEstadistica == 4):
                break
            
            
            if (opcionEstadistica == 1):
                x = ["Partidas ganadas","Partidas perdidas"]
                y = [self._datosJugadores[nombre][0], self._datosJugadores[nombre][1]]
                grafico.bar(x, y)
                grafico.title("Informacion del juego")
                grafico.show()
            elif (opcionEstadistica == 2):
                x = ["Fácil","Medio", "Difícil"]
                y = [self._datosJugadores[nombre][2], 
                     self._datosJugadores[nombre][3], 
                     self._datosJugadores[nombre][4]]
                grafico.bar(x, y)
                grafico.title("Partidas ganadas por dificultad del nivel")
                grafico.show()
            else :
                print("\nRanking de Jugadores con el número de partidas ganadas")
                aux = dict(sorted(self._datosJugadores.items(), key=operator.itemgetter(1), reverse=True))
                i=1
                for n in aux:
                   print(str(i) + ". " + str(n) + " " + str(aux[n][0])) 
                   i+=1
    
    
if __name__ == '__main__':
    juego = Menu()
    juego.menu()


